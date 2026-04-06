from __future__ import annotations

import shutil
from datetime import datetime
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = ROOT / "cz_rostliny_rozsireny_dataset_v6_jadro_bezne_trvanlive.xlsx"

SOURCE_FIELDS = ("zdroj_id", "nazev", "url", "poznamka")
STARTER_FIELDS = (
    "record_id",
    "vedecky_nazev",
    "cesky_nazev",
    "status_v_CR",
    "cast_rostliny",
    "domena",
    "poddomena",
    "status_znalosti",
    "aplikovatelnost_v_CR",
    "typicke_lokality_v_CR",
    "obdobi_ziskani",
    "fenologicka_faze",
    "zpusob_pripravy_nebo_vyuziti",
    "cilovy_efekt",
    "chutovy_profil",
    "vonny_profil",
    "palivovy_potencial",
    "hlavni_rizika",
    "kontraindikace_interakce",
    "legalita_poznamka_CR",
    "dukaznost_typ",
    "dukaznost_skore",
    "zdroj_id_1",
    "zdroj_id_2",
    "poznamka",
)
TRVANLIVE_FIELDS = (
    "record_id",
    "cesky_nazev",
    "vedecky_nazev",
    "domena",
    "poddomena",
    "cast_rostliny",
    "obdobi_ziskani",
    "typicke_lokality_v_CR",
    "forma_uchovani",
    "orientacni_trvanlivost",
    "poznamka_k_skladovani",
)


def rows(fields: tuple[str, ...], values: list[tuple[str, ...]]) -> list[dict[str, str]]:
    return [dict(zip(fields, item)) for item in values]


SOURCE_ROWS = rows(
    SOURCE_FIELDS,
    [
        ("S96", "PFAF: Tussilago farfara", "https://pfaf.org/user/Plant.aspx?LatinName=Tussilago+farfara", "Kurátorovaný zdroj pro časné jarní květní využití podbělu a opatrné použití na čaj se silným upozorněním na pyrrolizidinové alkaloidy."),
        ("S97", "PFAF: Primula veris", "https://pfaf.org/user/Plant.aspx?LatinName=Primula_veris", "Kurátorovaný zdroj pro prvosenku jarní: časné listy, květy, konzervy i jarní květní nápoje."),
        ("S98", "PFAF: Primula elatior", "https://pfaf.org/User/Plant.aspx?LatinName=Primula+elatior", "Kurátorovaný zdroj pro prvosenku vyšší jako velmi ranou lesní listovou zeleninu."),
        ("S99", "PFAF: Lamium purpureum", "https://pfaf.org/user/Plant.aspx?LatinName=Lamium_purpureum", "Kurátorovaný zdroj pro hluchavku nachovou jako nenápadnou, ale dobře dostupnou předjarní zeleň."),
        ("S100", "PFAF: Lamium album", "https://pfaf.org/user/plant.aspx?latinname=Lamium+album", "Kurátorovaný zdroj pro hluchavku bílou: mladé listy i sušené květy na jemný čaj."),
        ("S101", "PFAF: Alliaria petiolata", "https://pfaf.org/user/Plant.aspx?LatinName=alliaria+petiolata", "Kurátorovaný zdroj pro česnáček lékařský včetně listů, květů a mladých šešulek."),
        ("S102", "PFAF: Robinia pseudoacacia", "https://pfaf.org/user/Plant.aspx?LatinName=robinia_pseudoacacia", "Kurátorovaný zdroj pro květy akátu/trnovníku akátu; zároveň zdůrazňuje nutnost opatrnosti, protože ostatní části rostliny jsou toxické."),
        ("S103", "PFAF: Humulus lupulus", "https://pfaf.org/user/plant.aspx?latinname=Humulus+lupulus", "Kurátorovaný zdroj pro jarní chmelové výhonky a listy, využitelné jen v krátkém květnovém okně."),
        ("S104", "PFAF: Epilobium angustifolium", "https://pfaf.org/user/Plant.aspx?LatinName=Epilobium+angustifolium", "Kurátorovaný zdroj pro vrbovku úzkolistou jako jarní výhonkovou zeleninu i čajovou rostlinu."),
        ("S105", "UAF Extension: Fireweed", "https://www.uaf.edu/ces/publications/database/food/files/pdfs/FNH-00106-Fireweed-02-23-24.pdf", "Aljašský extension materiál s praktickými recepty na výhonky, želé, čaj a ledničkové pickles z vrbovky úzkolisté."),
        ("S106", "Luke / SPECICROP: fermentation of fireweed", "https://www.agrowebcee.net/uploads/media/luke-luobio_72_2016.pdf", "Finsko-ruský technický materiál k tradiční i poloprůmyslové fermentaci Epilobium angustifolium pro Ivan čaj."),
        ("S107", "PFAF: Silene vulgaris", "https://pfaf.org/user/Plant.aspx?LatinName=Silene_vulgaris", "Kurátorovaný zdroj pro silenku nadmutou jako výrazně ne-mainstreamovou evropskou jarní zeleninu."),
        ("S108", "PFAF: Tragopogon pratensis", "https://pfaf.org/user/Plant.aspx?LatinName=Tragopogon+pratensis", "Kurátorovaný zdroj pro kozí bradu luční včetně stonků a poupat podávaných jako 'divoký chřest'."),
        ("S109", "PFAF: Heracleum sphondylium", "https://pfaf.org/user/Plant.aspx?LatinName=Heracleum+Sphondylium", "Kurátorovaný zdroj pro bolševník obecný; použitelné jen s výrazným varováním kvůli fototoxicitě a záměně s jinými miříkovitými."),
    ],
)

STARTER_ROWS = rows(
    STARTER_FIELDS,
    [
        ("R257", "Tussilago farfara", "podběl lékařský", "původní, běžný", "květní poupata a mladé květy", "potrava", "časná květní poupata a květy do salátu / krátce tepelně", "téměř zapomenuté", "střední", "jílovité náspy, okraje cest, výsypky, rumiště, vlhčí svahy a narušené půdy", "III–IV", "čerstvě vyrašené květní stvoly a polootvřená poupata před plným rozkvětem", "Sbírat jen malé množství mladých květních poupat a čerstvých květů, použít je syrové jako aromatický detail do jarní směsi nebo je krátce tepelně upravit.", "Velmi raná, aromatická jarní květní jedlost, která posouvá sběr až do opravdu předjarního období.", "lehce anýzová, nahořkle aromatická", "jemně lékořicový, květní", "neaplikovatelné", "Rostlina obsahuje pyrrolizidinové alkaloidy; nejde o každodenní potravinu a sběr z krajnic nebo znečištěných náspů je nevhodný.", "Nevhodné v těhotenství, při kojení a pro malé děti; nepoužívat dlouhodobě ani ve velkém množství.", "Sbírat jen malé množství mimo chráněné lokality a mimo dopravně či chemicky zatížená místa.", "kurátorovaný botanický zdroj", "D", "S96", "", "Březnová květní vlna záměrně přidává i málo využívané předjarní druhy s opatrným rámováním rizik."),
        ("R258", "Tussilago farfara", "podběl lékařský", "původní, běžný", "listy a květy", "pití", "sušené listy a květy na občasný nálev", "téměř zapomenuté", "nízká až střední", "jílovité náspy, okraje cest, výsypky, rumiště, vlhčí svahy a narušené půdy", "III–V + sušení", "časné květy a později mladší listy sbírané jen z čistých míst", "Květy a mladší listy šetrně usušit a pak je používat jen střídmě a občas do lehkého bylinného nálevu.", "Historizující jarní čajová surovina s jemně lékořicovým profilem, ale nízkou praktickou prioritou kvůli rizikům.", "bylinná, lehce nahořklá, s nádechem lékořice", "jemně květní a bylinný", "neaplikovatelné", "Kvůli pyrrolizidinovým alkaloidům má dávat smysl jen občasné a střídmé použití; ne pro děti a ne dlouhodobě.", "Nevhodné v těhotenství, při kojení, u malých dětí a pro dlouhodobé pravidelné pití.", "Sbírat jen v malém, mimo chráněná území a mimo dopravně či chemicky zatížené lokality.", "kurátorovaný botanický zdroj", "D", "S96", "", "Zařazeno jako březnový trvanlivější relikt tradičního využití, ale vědomě mimo doporučený výběr."),
        ("R259", "Primula elatior", "prvosenka vyšší", "původní, lokálně běžná", "mladé listy", "potrava", "časné listy syrové / do polévky", "téměř zapomenuté", "střední", "vlhčí listnaté lesy, stinné lemy, humózní svahy a starší parky", "III–IV", "měkké přízemní listy ještě před hrubnutím a před plným rozkvětem", "Sbírat jen pár mladých listů z hojnějších trsů, použít je syrové do směsi nebo krátce přidat do jarní polévky.", "Velmi raná a nenápadná lesní listová zelenina pro uživatele, kteří chtějí jít hlouběji než po běžných plevelech.", "jemná, měkká, nevýrazně zelená", "slabě listový", "neaplikovatelné", "Sbírat jen tam, kde je druh opravdu hojný; nejde o rostlinu pro velkoobjemový sběr.", "", "V lučních a chráněných lokalitách být zdrženlivý a sbírat jen malé množství z míst, kde je sběr legální.", "kurátorovaný botanický zdroj", "D", "S98", "", "Březnové rozšíření směrem k velmi časným listům stinných stanovišť."),
        ("R260", "Primula veris", "prvosenka jarní", "původní, lokálně běžná", "mladé listy", "potrava", "časné listy syrové / do polévky / jako čajová náhražka", "téměř zapomenuté", "střední", "světlejší louky, remízky, okraje lesů, suché až mezofilní trávníky", "III–IV", "nejmladší listy v přízemní růžici před plným kvetením", "Mladé listy přidávat v malém množství do směsí nebo do jarní polévky; lze je i sušit a používat jako jednoduchou čajovou náhražku.", "Časná jarní listová surovina, která dává smysl hlavně pro rozšíření repertoáru o opravdu časné luční druhy.", "spíše jemná, lehce bylinná", "slabý zelený", "neaplikovatelné", "Listy nejsou chuťově výjimečné, takže sběr má smysl jen střídmý; nepustošit květné louky kvůli hrsti listů.", "U citlivých osob opatrnost na saponiny a při souběžném užívání léků na srážlivost.", "Sbírat jen střídmě a tam, kde je druh na lokalitě hojný a sběr je legální.", "kurátorovaný botanický zdroj", "D", "S97", "", "Březnová vlna se rozšiřuje i o časně sbíratelné prvosenkové listy."),
        ("R261", "Primula veris", "prvosenka jarní", "původní, lokálně běžná", "květy", "potrava", "květy do salátu / ozdoby / konzervy", "méně známé", "střední", "světlejší louky, remízky, okraje lesů, suché až mezofilní trávníky", "IV–V", "čerstvě rozvinuté voňavé květy", "Květy používat čerstvé jako jedlou ozdobu nebo je zpracovat do menší domácí konzervy, případně do tradičněji laděného jarního květního nápoje.", "Jemná květní jarní surovina s přesahem do domácích sladších konzerv a tradičnějších květních nápojů.", "jemně květní, lehce medová", "výrazněji jarně květní", "neaplikovatelné", "Netrhat ve velkém celé květy z málo početných lokalit; sběr má být spíš symbolický než objemový.", "U citlivých osob opatrnost na saponiny; při užívání antikoagulancií a v těhotenství konzervativní přístup.", "Sbírat jen malé množství z hojných porostů a vyhnout se chráněným nebo citlivým lokalitám.", "kurátorovaný botanický zdroj", "D", "S97", "", "Přemosťuje březen do května a dává datasetu i jemnější jarní květní konzervové směry."),
        ("R262", "Lamium purpureum", "hluchavka nachová", "původní, velmi běžná", "mladé listy a vrcholky", "potrava", "mladé vrcholky syrové / krátce vařené", "méně známé", "vysoká", "zahrady, záhony, okraje cest, pole, ruderální místa, spáry a městské kouty", "III–V", "měkké vrcholky před zhrubnutím a před plným odkvětem", "Vrcholky používat syrové po menším množství do jarních směsí nebo je krátce podusit s další zelení jako městsky dostupnou listovou složku.", "Nenápadná a snadno dostupná jarní zeleň z mikrostanovišť, která lidé obvykle míjejí bez povšimnutí.", "jemná, lehce zelená, mírně zemitá", "slabě bylinný", "neaplikovatelné", "Sbírat jen z opravdu čistých míst mimo psí zóny, krajnice a chemicky ošetřené záhony.", "", "Velmi běžný druh; prakticky nejbezpečnější je sběr ze zahrad nebo čistých městských mikrostanovišť.", "kurátorovaný botanický zdroj", "D", "S99", "", "Březnově-dubnová vlna přidává i málo zmiňované, ale velmi dostupné městské a zahradní druhy."),
        ("R263", "Lamium album", "hluchavka bílá", "původní, běžná", "mladé listy", "potrava", "mladé listy syrové / vařené", "méně známé", "střední až vysoká", "živé ploty, okraje cest, křoviny, lesní lemy, zahrady, ruderální místa", "III–V", "nejmladší listy a měkké vrcholky před zhrubnutím", "Mladé listy použít syrové do směsí nebo je vařit spolu s další jarní zelení; samostatně jsou spíš jemné než výrazné.", "Další tichý, ale použitelný jarní druh z živých plotů a lemů, vhodný hlavně do směsných sběrů.", "jemná, lehce travnatá", "slabě bylinný", "neaplikovatelné", "Sbírat z čistých míst a nesahat po listech z okrasných nebo chemicky ošetřených výsadeb.", "", "Sbírat střídmě a hlavně z běžných, snadno přístupných a čistých porostů.", "kurátorovaný botanický zdroj", "D", "S100", "", "Posiluje březnový a dubnový repertoár o další málo doceňovanou hluchavku."),
        ("R264", "Lamium album", "hluchavka bílá", "původní, běžná", "květy", "pití", "sušené květy na jemný čaj", "téměř zapomenuté", "střední", "živé ploty, křoviny, lemy cest, zahrady a ruderální okraje", "V–IX + sušení", "čerstvě rozvité bílé květy sbírané za sucha", "Květy šetrně usušit a používat na jemný bylinný čaj nebo jako zjemňující složku domácích směsí.", "Lehká a poměrně příjemná květní čajovina z velmi běžné, ale přehlížené rostliny.", "jemná, lehce nasládlá, květní", "slabě medový", "neaplikovatelné", "Sbírat jen suché, čisté květy; nezapařit je při sušení a brát je spíš jako jemnou čajovou surovinu než jako silnou bylinu.", "", "Běžný druh; vhodný hlavně ze zahradních a křovinných porostů bez chemického ošetření.", "kurátorovaný botanický zdroj", "D", "S100", "", "Květnová vlna vědomě přidává i tiché čajové suroviny, které nejsou v moderním foragingu moc vidět."),
        ("R265", "Alliaria petiolata", "česnáček lékařský", "původní, běžný", "mladé listy", "potrava", "časné listy do salátu / jako dochucení", "méně známé", "vysoká", "křoviny, lesní lemy, stinnější zídky, staré sady, vlhčí polostinná místa", "III–V", "mladé listy před plným kvetením", "Listy používat syrové do salátů a pomazánek nebo po malém množství do vařených jídel jako česnekově-hořčičné dochucení.", "Brzy dostupná kořenná listová bylina, která dobře překlene období před česnekem medvědím a dalšími výraznějšími zeleněmi.", "česnekově-hořčičná, lehce štiplavá", "česnekový, brukvovitý", "neaplikovatelné", "Jako brukvovitá může ve větším množství dráždit citlivější trávení; sbírat mimo znečištěná městská stinná místa.", "U citlivějších osob opatrnost při vyšším množství brukvovitých.", "Běžný druh; sběr v malém z běžných porostů je prakticky dostupný, ale vždy mimo chráněné lokality a znečištěná místa.", "kurátorovaný botanický zdroj", "D", "S101", "", "Přidává do března a dubna silnější chuťový most mezi listovou zelení a kořením."),
        ("R266", "Alliaria petiolata", "česnáček lékařský", "původní, běžný", "květy a mladé šešulky", "potrava", "květy a mladé šešulky syrové jako kořeněná složka", "téměř zapomenuté", "střední až vysoká", "křoviny, lesní lemy, starší parky, zídky a stinná rumiště", "V–VI", "otevřené květy a velmi mladé měkké šešulky", "Květy a ještě měkké mladé šešulky používat syrové po malém množství jako pikantní, česnekově-hořčičný detail do jídel.", "Květnová nadstavba česnáčku, která rozšiřuje sběr i po odeznění nejlepších listů.", "pikantní, česnekově-hořčičný", "výrazně brukvovitý", "neaplikovatelné", "Sbírat jen mladé, měkké šešulky; starší části rychle hrubnou a chuť bývá tvrdší.", "U citlivějších osob opatrnost při vyšším množství brukvovitých.", "Běžný druh; sbírat střídmě a jen z čistých míst.", "kurátorovaný botanický zdroj", "D", "S101", "", "Květnová vlna doplňuje i méně zmiňované pozdější části česnáčku."),
        ("R267", "Robinia pseudoacacia", "trnovník akát", "nepůvodní, zdomácnělý a běžný", "květy", "potrava", "květy vařené / nápoj / zavařenina", "méně známé", "střední", "teplejší stráně, aleje, remízky, suché meze, městské a příměstské výsadby", "V–VI", "čerstvě rozvinuté voňavé květní hrozny", "Sbírat jen samotné květy bez zelených částí a používat je vařené do sladších jídel, do domácí zavařeniny nebo na voňavý květní nápoj.", "Výrazně aromatická květnová surovina s potenciálem pro nápojové a sladké uchování.", "sladce květní, lehce vanilkový", "silně medově květní", "neaplikovatelné", "Jedlé jsou jen květy; listy, kůra, semena a další části jsou toxické. Nesbírat z prašných ulic ani stromů po chemickém zásahu.", "Používat jen květy a při první práci s druhem začít opatrně malým množstvím.", "Běžný zdomácnělý druh; sběr jen z čistých míst a bez poškozování výhonů či stromů.", "kurátorovaný botanický zdroj", "D", "S102", "", "Květnová vlna přidává i silně voňavé květní využití s velkým důrazem na to, že jedlé jsou opravdu jen květy."),
        ("R268", "Humulus lupulus", "chmel otáčivý", "původní, běžný", "mladé výhonky a listy", "potrava", "jarní výhonky a listy vařené / syrové po menším množství", "téměř zapomenuté", "střední", "křoviny, břehy, nivy, remízky, okraje lesů a staré ploty", "IV–V", "nejmladší špičky výhonů a měkké listy před koncem května", "Mladé výhonky krátce vařit nebo podusit jako jarní zeleninu; nejmladší listy lze po malém množství přidat i syrové do směsí.", "Krátké a hodnotné květnové okno pro výraznější divoké výhonky, které většina lidí spojuje spíš s pivem než s jídlem.", "zelený, lehce chmelový, specifický", "pryskyřičně bylinný", "neaplikovatelné", "Používat jen opravdu mladé části před koncem května; starší rychle hrubnou a ztrácejí kuchyňskou hodnotu.", "", "Sbírat jednotlivé špičky z hojných porostů bez strhávání celých lian.", "kurátorovaný botanický zdroj", "D", "S103", "", "Květnová vlna rozšiřuje výhonkový sběr i na druhy, které jsou dnes v kuchyňském použití skoro neviditelné."),
        ("R269", "Epilobium angustifolium", "vrbovka úzkolistá", "původní, běžná", "mladé výhonky, listy a poupata", "potrava", "jarní výhonky syrové / vařené / ledničkové pickles", "téměř zapomenuté", "střední", "lesní paseky, okraje cest v horách a podhůří, světliny, náspy, disturbované lesní okraje", "V–VI", "mladé šťavnaté výhonky, jemné listy a neotevřená poupata před plným květem", "Mladé výhonky a poupata používat syrové po menším množství, krátce je napařit nebo je naložit do ledničkových pickles.", "Květnová a raně letní výhonková zelenina s překvapivě širokým rozsahem od čerstvého použití po jemnější pickles.", "jemně zelený, někdy lehce peprný", "čerstvě bylinný", "neaplikovatelné", "Pozdější listy a stonky rychle tuhnou a hořknou; sběr má smysl jen opravdu časný a z čistých lokalit.", "", "Sbírat střídmě jen vrcholové části z běžných porostů a neničit celé plochy mladých výhonů.", "kurátorovaný botanický + extension zdroj", "D", "S104", "S105", "Květnová vlna přidává i rostlinu, která umí spojit čerstvý sběr s praktickým uchováním."),
        ("R270", "Epilobium angustifolium", "vrbovka úzkolistá", "původní, běžná", "mladé listy a vrcholky", "pití", "fermentované a sušené listy na Ivan čaj", "známé hlavně v zahraničí", "střední", "světliny, paseky, náspy, horské a podhorské disturbované lokality", "V–VIII + fermentace + sušení", "mladé listy a vrcholky v době před květem až na začátku nasazování poupat", "Listy nebo měkké vrcholky nechat zavadnout, narušit pletiva, krátce fermentovat a následně dosušit jako nečerný bylinný 'Ivan čaj'.", "Tradičněji laděná fermentovaná čajová surovina, známější na severu a východě Evropy než u nás.", "hladký, tmavší bylinný, lehce ovocný", "jemně fermentovaný, čajový", "neaplikovatelné", "Bez správného zavadnutí, fermentace a dosušení může surovina plesnivět; sbírat jen zdravé listy bez prachu.", "Jako u silnějších bylinných čajů začínat rozumným množstvím.", "Sbírat jen z hojných porostů a neodřezávat celé plochy vrcholových částí.", "kurátorovaný botanický + technologický zdroj", "D", "S104", "S106", "Květnové rozšíření vědomě přidává i zahraničněji známý fermentační směr, který odpovídá požadavku pronikat do tajů."),
        ("R271", "Silene vulgaris", "silenka nadmutá", "původní, běžná lokálně", "mladé výhonky a listy", "potrava", "mladé výhonky syrové / vařené před kvetením", "známé hlavně v zahraničí", "střední", "suché meze, stráně, louky, okraje cest a kamenitější trávníky", "V–VI", "krátké mladé výhonky a listy před rozkvětem", "Mladé výhonky jíst po menším množství syrové nebo je krátce povařit; při větší hořkosti je vhodné je předem blanšírovat.", "Velmi ne-mainstreamová evropská jarní zelenina, která je u nás skoro neviditelná, ale jinde je ceněná.", "sladce zelený s lehkou hořčinou", "slabě bylinno-zelený", "neaplikovatelné", "Používat jen mladé části před květem a ne ve velkých objemech kvůli obsahu saponinů.", "Nevhodné ve velkých dávkách; při citlivém trávení začínat malým množstvím a preferovat tepelnou úpravu.", "Sbírat střídmě jen z hojných porostů a ne z lokalit podezřelých na těžké kovy či jinou zátěž.", "kurátorovaný botanický zdroj", "D", "S107", "", "Květnová vlna úmyslně otevírá i druhy známé spíš v zahraničních regionálních kuchyních."),
        ("R272", "Tragopogon pratensis", "kozí brada luční", "původní, běžná", "mladé stonky, poupata a listy", "potrava", "květní stonky a poupata vařené jako 'divoký chřest'", "téměř zapomenuté", "střední", "louky, meze, pastviny, náspy a světlejší ruderální okraje", "V–VI", "mladé květní stonky a poupata před otevřením květů", "Květní stonky s poupaty krátce povařit nebo podusit a podávat podobně jako divoký chřest; listy a mladé výhonky lze přidat i do polévky.", "Květnový 'chřestový' detail z luční rostliny, která se dnes téměř vůbec nečte jako kuchyňská.", "jemně nasládlý, zeleninový", "slabě bylinný", "neaplikovatelné", "Používat jen mladé stonky a poupata; starší části rychle dřevnatí a kvalita jde dolů.", "", "Sbírat střídmě z běžných lučních porostů a nepřerušovat celé populace v době květu.", "kurátorovaný botanický zdroj", "D", "S108", "", "Květnové rozšíření posouvá dataset i směrem k lučním stonkovým a poupatovým jedlostem."),
        ("R273", "Heracleum sphondylium", "bolševník obecný", "původní, běžný", "mladé stonky, řapíky a výhonky", "potrava", "mladé stonky a výhonky vařené jako zelenina", "téměř zapomenuté", "nízká až střední", "vlhčí louky, příkopy, lemy cest, křoviny a lesní okraje", "V", "úplně mladé výhonky a stonky těsně po vyrašení, před kvetením", "Použitelné jsou jen mladé výhonky a stonky, ideálně po oloupání a tepelné úpravě; starší části rychle ztrácejí kvalitu.", "Vysoce niche květnová zelenina pro opravdu pokročilé sběrače, historicky ceněná, dnes skoro neznámá.", "výrazně zeleninový, lehce aromatický", "miříkovitý, kořenitý", "neaplikovatelné", "Fototoxicita a riziko záměny s jinými miříkovitými včetně nebezpečných druhů; bez naprosté jistoty určování a ochrany kůže nesbírat.", "Nevhodné pro začátečníky; citlivá kůže může reagovat i na mízu.", "Sbírat jen malé množství z běžných porostů a pouze tam, kde je sběr legální a nehrozí poškození citlivého stanoviště.", "kurátorovaný botanický zdroj", "D", "S109", "", "Zařazeno záměrně jako 'tajná' květnová rostlina, ale s nejostřejším bezpečnostním rámcem z celé vlny."),
    ],
)

TRVANLIVE_ROWS = rows(
    TRVANLIVE_FIELDS,
    [
        ("R258", "podběl lékařský", "Tussilago farfara", "pití", "sušené listy a květy na občasný nálev", "listy a květy", "III–V + sušení", "jílovité náspy, okraje cest, výsypky, rumiště, vlhčí svahy a narušené půdy", "sušení na čaj", "cca 6–12 měsíců", "Sušit rychle v tenké vrstvě, skladovat v suchu a temnu. Kvůli alkaloidům jen střídmé a občasné použití."),
        ("R261", "prvosenka jarní", "Primula veris", "potrava", "květy do salátu / ozdoby / konzervy", "květy", "IV–V", "světlejší louky, remízky, okraje lesů, suché až mezofilní trávníky", "zavařenina / květní konzerva", "cca 6–12 měsíců", "Zpracovat jen čisté květy bez stopek; po otevření uchovávat v chladu."),
        ("R264", "hluchavka bílá", "Lamium album", "pití", "sušené květy na jemný čaj", "květy", "V–IX + sušení", "živé ploty, křoviny, lemy cest, zahrady a ruderální okraje", "sušení na čaj", "cca 6–12 měsíců", "Sušit v tenké vrstvě a chránit před zapařením. Uchovávat v uzavřené nádobě a temnu."),
        ("R267", "trnovník akát", "Robinia pseudoacacia", "potrava", "květy vařené / nápoj / zavařenina", "květy", "V–VI", "teplejší stráně, aleje, remízky, suché meze, městské a příměstské výsadby", "sirup / zavařenina", "cca 6–12 měsíců", "Zpracovat jen samotné květy bez zelených částí; po otevření skladovat v chladu a vždy hlídat čistotu suroviny."),
        ("R270", "vrbovka úzkolistá", "Epilobium angustifolium", "pití", "fermentované a sušené listy na Ivan čaj", "mladé listy a vrcholky", "V–VIII + fermentace + sušení", "světliny, paseky, náspy, horské a podhorské disturbované lokality", "fermentace + sušení", "cca 6–18 měsíců", "Po fermentaci důkladně dosušit, jinak hrozí plesnivění. Uchovávat v suchu, temnu a těsné nádobě."),
    ],
)


def backup_workbook(path: Path) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = path.with_name(f"{path.stem}.backup_{timestamp}{path.suffix}")
    shutil.copy2(path, backup_path)
    return backup_path


def header_index_map(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, int]:
    headers = [cell.value for cell in ws[1]]
    return {str(value): index + 1 for index, value in enumerate(headers) if value is not None}


def upsert_rows(ws: openpyxl.worksheet.worksheet.Worksheet, key_field: str, payloads: list[dict[str, str]]) -> tuple[int, int]:
    headers = header_index_map(ws)
    if key_field not in headers:
        raise KeyError(f"Missing key field {key_field} in sheet {ws.title}")

    existing: dict[str, int] = {}
    for row_index in range(2, ws.max_row + 1):
        value = ws.cell(row=row_index, column=headers[key_field]).value
        if value is not None:
            existing[str(value)] = row_index

    inserted = 0
    updated = 0
    for payload in payloads:
        key_value = str(payload[key_field])
        row_index = existing.get(key_value)
        if row_index is None:
            row_index = ws.max_row + 1
            existing[key_value] = row_index
            inserted += 1
        else:
            updated += 1

        for field, value in payload.items():
            column = headers.get(field)
            if column is not None:
                ws.cell(row=row_index, column=column).value = value

    return inserted, updated


def main() -> None:
    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(f"Workbook not found: {WORKBOOK_PATH}")

    backup_path = backup_workbook(WORKBOOK_PATH)
    workbook = openpyxl.load_workbook(WORKBOOK_PATH)

    zdroje_inserted, zdroje_updated = upsert_rows(workbook["Zdroje"], "zdroj_id", SOURCE_ROWS)
    starter_inserted, starter_updated = upsert_rows(workbook["Starter_dataset"], "record_id", STARTER_ROWS)
    trvanlive_inserted, trvanlive_updated = upsert_rows(workbook["Trvanlive_1m_plus"], "record_id", TRVANLIVE_ROWS)

    workbook.save(WORKBOOK_PATH)

    print(f"Workbook backup: {backup_path}")
    print(f"Zdroje inserted={zdroje_inserted} updated={zdroje_updated}")
    print(f"Starter_dataset inserted={starter_inserted} updated={starter_updated}")
    print(f"Trvanlive_1m_plus inserted={trvanlive_inserted} updated={trvanlive_updated}")


if __name__ == "__main__":
    main()
